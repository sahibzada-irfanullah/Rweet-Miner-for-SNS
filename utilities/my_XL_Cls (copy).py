
# coding: utf-8

# In[38]:

import pandas as pd
import xlsxwriter
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook

# fname = '/home/sahibzada/Desktop/ipythonNB/ThImp/results/pandas (copy).xlsx'


class XL_Results_writing:
  
  fileName = None
  
  def __init__(self, fileName=None):
    if fileName is None:
      print("Please, provide complete path to the file.")
    else:
      self.fileName = fileName

  def generate_resultantWorkSheet(self, data_column, result_sheet_name):
    '''
    -generate features vs classifiers dataframe from worksheet of the workbook and save to new worksheet
    at the end of this book
    +take fileName i.e., workbook, names of the column which is to extracted from each worksheet and
    name of sheet in which resultant table would be stored
    '''
    print("Generating resultant sheet")
    self.remove_extraWorkSheet('Sheet1')
    book = load_workbook(self.fileName)
    df = pd.read_excel(self.fileName, sheetname=None, skiprows=3)
    combined = pd.DataFrame()
    columns_names = []
    sheet_names = []
    frames = []
    frequency_list = []
    for ws in book.worksheets:
      #     print(ws['D2'].value)
      if ws.title != result_sheet_name:
        columns_names.append(ws['D2'].value)
        frequency_list.append(ws['E2'].value)
        frames.append(df[ws.title][data_column])
        temp_sheet = ws.title
    combined = pd.concat(frames, axis=1)
    combined.columns = columns_names
    combined.insert(0, df[temp_sheet].columns[1], df[temp_sheet].iloc[:, 1])
    combined.set_index(df[temp_sheet].columns[1], inplace=True)
    #     print(combined)
    self.save_resultsToExcel(combined, result_sheet_name, 'Phrases\' combinations')
    book.close()
    self.save_freqs2resultSheet(frequency_list, result_sheet_name)

  def save_freqs2resultSheet(self, frequency_list, result_sheet):
    book = load_workbook(self.fileName)
    res_sheet = book.get_sheet_by_name(result_sheet)
    cell_list = ['C3', 'D3', 'E3', 'F3', 'G3', 'H3']
    for cell, value in zip(cell_list, frequency_list):
      res_sheet[cell] = value
    book.save(self.fileName)
    book.close()

  def show_allSheets(self):
    book = load_workbook(self.fileName)
    for ws in book.worksheets:
          print(ws.title)
    book.close()
    
  def remove_extraWorkSheet(self, sheet_name):
    '''
    -remove extra sheet from worksheet
    + takes filename which contans sheet and sheet name which should be deleted
    '''
    sheet_name = sheet_name
    book = load_workbook(self.fileName)
    for ws in book.worksheets:
#           print(ws)
          if ws.title == sheet_name:
            book.remove_sheet(book.get_sheet_by_name(sheet_name))
            book.save(self.fileName)
            book.close()
            print("Extra sheet \'"+ sheet_name + "\' is removed")

  def ls_ToDf(self,ls1, ls2, ls3, ls4, ls5, ls6):
    '''
    -convert list type into DataFrame and add columns names for creating labelled table of dataframe type
    + takes three lists of arguments and convert it Dataframe row wise order
    '''
    ls_digit1 = list(map(float, ls1))
    ls_digit2 = list(map(float, ls2))
    ls_digit3 = list(map(float, ls3))
    ls_digit4 = list(map(float, ls4))
    ls_digit5 = list(map(float, ls5))
    ls_digit6 = list(map(float, ls6))
    res = list()
    res.append(ls_digit1)
    res.append(ls_digit2)
    res.append(ls_digit3)
    res.append(ls_digit4)
    res.append(ls_digit5)
    res.append(ls_digit6)
    df = pd.DataFrame(res, columns=['Accuracy', 'Precision', 'Recall', 'F1-Measure'])
  #   df['Cassifire']=['Naive bayes', 'Logistic Regression', 'SVM', 'Random Forest', 'NLP', 'Gradient Boosting']
  #   df = df[['Cassifire','Accuracy', 'Precision', 'Recall', 'F1-Measure']]
    df['Classifier']=['Naive Bayes', 'Logisitic Regression', 'SVM', 'Random Forest', 'Gradient Boosting', 'NLP']
    df = df[['Classifier','Accuracy', 'Precision', 'Recall', 'F1-Measure']]
    df.set_index('Classifier', inplace = True)
    return df

  def save_resultsToExcel(self, df, sheet_name, feature, freq=''):
    '''
    -save dataframe to a sheet and statically write something in one cell
    +takes dataframe handler, sheet_name in which that dataframe to be stored, feature for writing 
    in a single cell and fileName is the workbook-excel file
    '''
  #   workbook   = xlsxwriter.Workbook('filename.xlsx')
    book = load_workbook(self.fileName)
    writer = pd.ExcelWriter(self.fileName, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer,startcol=1 ,startrow=3, sheet_name=sheet_name)
    sheetname = book.get_sheet_by_name(sheet_name)
  #   book.remove_sheet(book.get_sheet_by_name("Sheet1"))
  #   sheetname.merge_cells('C2:H2')
    sheetname['D2'] = feature
    sheetname['E2'] = freq
    writer.save()
    book.close()
    print("\nResults saved to the worksheet")

#
# xl = XL_Results_writing(fname)
# # xl.show_allSheets()
#
# ls1 = ['10.82', '73.38', '90.89', '13.00']
# ls2 = ['20.82', '76.88', '96.89', '23.00']
# ls3 = ['30.82', '66.38', '92.89', '33.60']
# df1 = xl.ls_ToDf(ls1, ls2, ls3)
# xl.save_resultsToExcel(df1, 'sheet-11','Uni grams')
#
#
# ls4 = ['40.82', '73.38', '90.89', '43.00']
# ls5 = ['50.82', '79.88', '68.89', '53.00']
# ls6 = ['60.82', '67.38', '92.89', '63.60']
# df2 = xl.ls_ToDf(ls4, ls5, ls6)
# xl.save_resultsToExcel(df2, 'sheet-12', 'Bi grams')
#
#
# ls7 = ['70.82', '73.38', '70.89', '73.00']
# ls8 = ['80.82', '76.88', '86.89', '83.00']
# ls9 = ['90.82', '76.38', '62.89', '93.60']
# df3 = xl.ls_ToDf(ls7, ls8, ls9)
# xl.save_resultsToExcel(df3, 'sheet-13', 'Tri grams')
# xl.generate_resultantWorkSheet('Accuracy', 'Results')
# # # xl.remove_extraWorkSheet('Sheet1')
#
#
#
# # In[29]:
#
# fname = '/home/sahibzada/Desktop/ipythonNB/ThImp/results/pandas (copy).xlsx'
# test = XL_Results_writing(fname)
# print(help())
#
#
# # ls1 = ['10.82', '73.38', '90.89', '13.00']
# # ls2 = ['20.82', '76.88', '96.89', '23.00']
# # ls3 = ['30.82', '66.38', '92.89', '33.60']
# # df1 = test.ls_ToDf(ls1, ls2, ls3)
# # test.save_resultsToExcel(df1, 'sheet-11','Uni grams')
#
#
# # ls4 = ['40.82', '73.38', '90.89', '43.00']
# # ls5 = ['50.82', '79.88', '68.89', '53.00']
# # ls6 = ['60.82', '67.38', '92.89', '63.60']
# # df2 = test.ls_ToDf(ls4, ls5, ls6)
# # test.save_resultsToExcel(df2, 'sheet-12', 'Bi grams')
#
#
# # ls7 = ['70.82', '73.38', '70.89', '73.00']
# # ls8 = ['80.82', '76.88', '86.89', '83.00']
# # ls9 = ['90.82', '76.38', '62.89', '93.60']
# # df3 = test.ls_ToDf(ls7, ls8, ls9)
# # test.save_resultsToExcel(df3, 'sheet-13', 'Tri grams')
# # test.remove_extraWorkSheet('Sheet1')
# # test.generate_resultantWorkSheet('Accuracy', 'Results')
# # test.generate_resultantWorkSheet('Accuracy', 'Resultant')
#
#
